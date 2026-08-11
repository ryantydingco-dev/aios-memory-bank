#!/usr/bin/env python3
"""Validate CA's local revenue tracker and generate daily/weekly operating views.

Local-only by design:
- reads one .xlsx workbook;
- writes Markdown views;
- makes no network calls;
- changes no workbook, CRM, QuickBooks, Mailchimp, or message state.

Usage:
  .venv/bin/python scripts/ca_revenue_ops.py validate --workbook path/to/book.xlsx
  .venv/bin/python scripts/ca_revenue_ops.py views --workbook path/to/book.xlsx \
      --output-dir outputs/revenue-ops-views --as-of 2026-07-28
"""

from __future__ import annotations

import argparse
import json
import math
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


SHEET_NAME = "Revenue Tracker"
HEADER_ROW = 6
DATA_START_ROW = 7

REQUIRED_COLUMNS = [
    "record_type",
    "opportunity_id",
    "account",
    "relationship",
    "customer_segment",
    "buyer_role",
    "offer_package",
    "product_category",
    "lead_source",
    "campaign_id",
    "first_touch_date",
    "stage",
    "quote_date",
    "quote_amount",
    "expected_order_value",
    "probability_override",
    "close_probability",
    "weighted_pipeline",
    "order_date",
    "order_value",
    "qb_customer_id",
    "qb_invoice_id",
    "reorder_flag",
    "next_action",
    "next_action_due",
    "next_action_status",
    "owner",
    "last_touch_date",
    "days_open",
    "loss_reason",
    "campaign_attribution",
    "marketing_permission",
    "approval_status",
    "system_of_record",
    "external_record_id",
    "notes",
    "data_quality_flag",
]

STAGE_PROBABILITY = {
    "identified": 0.05,
    "contact_ready": 0.10,
    "outreach_drafted": 0.10,
    "contacted": 0.10,
    "engaged": 0.20,
    "discovery": 0.35,
    "quote_requested": 0.45,
    "quote_sent": 0.60,
    "verbal_yes": 0.80,
    "proofing": 0.90,
    "won": 1.00,
    "nurture": 0.10,
    "lost": 0.00,
}

OPEN_STAGES = set(STAGE_PROBABILITY) - {"won", "lost"}
RELATIONSHIPS = {"existing_customer", "new_customer", "partner"}
SEGMENTS = {
    "recent_buyer",
    "seasonal_reorder",
    "high_value",
    "lapsed",
    "new_prospect",
    "other",
}
PERMISSIONS = {
    "subscribed",
    "unknown",
    "transactional_only",
    "unsubscribed",
    "do_not_contact",
    "not_applicable",
}
APPROVALS = {"draft", "needs_review", "approved", "hold", "not_applicable"}
SYSTEMS = {"local_tracker", "hubspot", "quickbooks", "mailchimp", "other"}
REORDER_FLAGS = {"yes", "no", "unknown"}
LOSS_REASONS = {
    "timing",
    "budget",
    "no_need",
    "incumbent_vendor",
    "unreachable",
    "not_decision_maker",
    "deadline_infeasible",
    "minimum_quantity",
    "product_fit",
    "price",
    "procurement",
    "service_issue",
    "duplicate_existing",
    "no_permission",
    "other",
}


@dataclass(frozen=True)
class Issue:
    severity: str
    row: int | None
    field: str
    message: str


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def parse_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value) if isinstance(value, float) else False:
            return None
        return float(value)
    text = clean_text(value).replace("$", "").replace(",", "")
    if not text or text.startswith("="):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text or text.startswith("="):
            return None
        try:
            return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
        except ValueError:
            try:
                return date.fromisoformat(text[:10])
            except ValueError:
                return None
    return None


def money(value: float | None) -> str:
    return f"${(value or 0):,.0f}"


def load_tracker(path: Path) -> tuple[list[dict[str, Any]], list[Issue]]:
    issues: list[Issue] = []
    if not path.is_file():
        return [], [Issue("error", None, "workbook", f"Workbook not found: {path}")]

    try:
        workbook = load_workbook(path, data_only=False, read_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl supplies varied error classes
        return [], [Issue("error", None, "workbook", f"Cannot open workbook: {exc}")]

    if SHEET_NAME not in workbook.sheetnames:
        workbook.close()
        return [], [
            Issue("error", None, "sheet", f"Required sheet is missing: {SHEET_NAME}")
        ]

    sheet = workbook[SHEET_NAME]
    headers = [clean_text(cell.value) for cell in sheet[HEADER_ROW]]
    while headers and not headers[-1]:
        headers.pop()

    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    extra = [column for column in headers if column and column not in REQUIRED_COLUMNS]
    if missing:
        issues.append(
            Issue("error", HEADER_ROW, "headers", f"Missing columns: {', '.join(missing)}")
        )
    if extra:
        issues.append(
            Issue("warning", HEADER_ROW, "headers", f"Unexpected columns: {', '.join(extra)}")
        )
    if issues and any(issue.severity == "error" for issue in issues):
        workbook.close()
        return [], issues

    rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=DATA_START_ROW, max_col=len(headers), values_only=True),
        start=DATA_START_ROW,
    ):
        record = dict(zip(headers, values))
        record["_row"] = row_number
        record_type = clean_text(record.get("record_type"))
        opportunity_id = clean_text(record.get("opportunity_id"))
        if not record_type and not opportunity_id:
            continue
        if record_type != "DATA":
            # Keep synthetic examples and notes out of operating metrics.
            if record_type not in {"SYNTHETIC_EXAMPLE", "NOTE"}:
                issues.append(
                    Issue("warning", row_number, "record_type", f"Unknown type: {record_type}")
                )
            continue
        rows.append(record)

    workbook.close()
    return rows, issues


def validate_rows(rows: Iterable[dict[str, Any]]) -> list[Issue]:
    issues: list[Issue] = []
    seen: dict[str, int] = {}

    for record in rows:
        row = int(record["_row"])
        opportunity_id = clean_text(record.get("opportunity_id"))
        account = clean_text(record.get("account"))
        stage = clean_text(record.get("stage"))
        relationship = clean_text(record.get("relationship"))
        segment = clean_text(record.get("customer_segment"))
        permission = clean_text(record.get("marketing_permission"))
        approval = clean_text(record.get("approval_status"))
        system = clean_text(record.get("system_of_record"))
        reorder = clean_text(record.get("reorder_flag"))
        loss_reason = clean_text(record.get("loss_reason"))

        if not opportunity_id:
            issues.append(Issue("error", row, "opportunity_id", "Required for DATA rows"))
        elif opportunity_id in seen:
            issues.append(
                Issue(
                    "error",
                    row,
                    "opportunity_id",
                    f"Duplicate of row {seen[opportunity_id]}: {opportunity_id}",
                )
            )
        else:
            seen[opportunity_id] = row

        for field, value in (
            ("account", account),
            ("lead_source", clean_text(record.get("lead_source"))),
            ("stage", stage),
            ("owner", clean_text(record.get("owner"))),
            ("system_of_record", system),
        ):
            if not value:
                issues.append(Issue("error", row, field, "Required for DATA rows"))

        if stage and stage not in STAGE_PROBABILITY:
            issues.append(Issue("error", row, "stage", f"Unknown stage: {stage}"))
        if relationship and relationship not in RELATIONSHIPS:
            issues.append(
                Issue("error", row, "relationship", f"Unknown relationship: {relationship}")
            )
        if segment and segment not in SEGMENTS:
            issues.append(
                Issue("error", row, "customer_segment", f"Unknown segment: {segment}")
            )
        if permission and permission not in PERMISSIONS:
            issues.append(
                Issue(
                    "error",
                    row,
                    "marketing_permission",
                    f"Unknown permission: {permission}",
                )
            )
        if approval and approval not in APPROVALS:
            issues.append(
                Issue("error", row, "approval_status", f"Unknown approval: {approval}")
            )
        if system and system not in SYSTEMS:
            issues.append(
                Issue("error", row, "system_of_record", f"Unknown system: {system}")
            )
        if reorder and reorder not in REORDER_FLAGS:
            issues.append(
                Issue("error", row, "reorder_flag", f"Unknown reorder flag: {reorder}")
            )

        probability_override = parse_number(record.get("probability_override"))
        if probability_override is not None and not 0 <= probability_override <= 1:
            issues.append(
                Issue(
                    "error",
                    row,
                    "probability_override",
                    "Must be a decimal between 0 and 1",
                )
            )

        for field in ("quote_amount", "expected_order_value", "order_value"):
            number = parse_number(record.get(field))
            if number is not None and number < 0:
                issues.append(Issue("error", row, field, "Cannot be negative"))

        if stage in OPEN_STAGES:
            if not clean_text(record.get("next_action")):
                issues.append(Issue("error", row, "next_action", "Open stage needs an action"))
            if not parse_date(record.get("next_action_due")):
                issues.append(
                    Issue("error", row, "next_action_due", "Open stage needs a valid due date")
                )

        if stage == "lost":
            if not loss_reason:
                issues.append(Issue("error", row, "loss_reason", "Lost stage needs a reason"))
            elif loss_reason not in LOSS_REASONS:
                issues.append(
                    Issue("error", row, "loss_reason", f"Unknown loss reason: {loss_reason}")
                )

        if stage == "won":
            if not parse_date(record.get("order_date")):
                issues.append(Issue("error", row, "order_date", "Won stage needs an order date"))
            if not (parse_number(record.get("order_value")) or 0) > 0:
                issues.append(
                    Issue("error", row, "order_value", "Won stage needs a positive order value")
                )
            if not clean_text(record.get("qb_invoice_id")):
                issues.append(
                    Issue(
                        "warning",
                        row,
                        "qb_invoice_id",
                        "Won row is not yet reconciled to a QuickBooks invoice",
                    )
                )

        lead_source = clean_text(record.get("lead_source"))
        campaign_id = clean_text(record.get("campaign_id"))
        attribution = clean_text(record.get("campaign_attribution"))
        if lead_source == "existing_customer_mailchimp" and permission != "subscribed":
            issues.append(
                Issue(
                    "error",
                    row,
                    "marketing_permission",
                    "Mailchimp-sourced row must be subscribed",
                )
            )
        if campaign_id and not attribution:
            issues.append(
                Issue(
                    "warning",
                    row,
                    "campaign_attribution",
                    "Campaign ID is present but attribution is blank",
                )
            )
        if approval == "approved" and permission in {
            "unknown",
            "transactional_only",
            "unsubscribed",
            "do_not_contact",
        }:
            issues.append(
                Issue(
                    "error",
                    row,
                    "approval_status",
                    f"Cannot approve marketing with permission={permission}",
                )
            )

    return issues


def validate_workbook(path: Path) -> tuple[list[dict[str, Any]], list[Issue]]:
    rows, issues = load_tracker(path)
    issues.extend(validate_rows(rows))
    return rows, issues


def effective_probability(record: dict[str, Any]) -> float:
    override = parse_number(record.get("probability_override"))
    if override is not None:
        return override
    return STAGE_PROBABILITY.get(clean_text(record.get("stage")), 0.0)


def expected_value(record: dict[str, Any]) -> float:
    return parse_number(record.get("expected_order_value")) or 0.0


def order_value(record: dict[str, Any]) -> float:
    return parse_number(record.get("order_value")) or 0.0


def row_label(record: dict[str, Any]) -> str:
    account = clean_text(record.get("account")) or "(missing account)"
    opportunity_id = clean_text(record.get("opportunity_id"))
    return f"{account} (`{opportunity_id}`)"


def daily_view(rows: list[dict[str, Any]], as_of: date) -> str:
    actions = []
    for record in rows:
        stage = clean_text(record.get("stage"))
        due = parse_date(record.get("next_action_due"))
        if stage not in OPEN_STAGES or not due or due > as_of + timedelta(days=7):
            continue
        actions.append((due, -expected_value(record), record))
    actions.sort(key=lambda item: (item[0], item[1], row_label(item[2]).lower()))

    overdue = sum(1 for due, _, _ in actions if due < as_of)
    due_today = sum(1 for due, _, _ in actions if due == as_of)
    open_pipeline = sum(
        expected_value(record)
        for record in rows
        if clean_text(record.get("stage")) in OPEN_STAGES
    )
    weighted = sum(
        expected_value(record) * effective_probability(record)
        for record in rows
        if clean_text(record.get("stage")) in OPEN_STAGES
    )

    lines = [
        f"# Revenue daily view — {as_of.isoformat()}",
        "",
        "> Local tracker view. No external action is performed.",
        "",
        f"- Open pipeline: **{money(open_pipeline)}**",
        f"- Weighted pipeline: **{money(weighted)}**",
        f"- Overdue actions: **{overdue}**",
        f"- Due today: **{due_today}**",
        "",
        "## Actions due through the next 7 days",
        "",
    ]
    if not actions:
        lines.append("_No dated actions due._")
    else:
        lines.extend(
            [
                "| Due | Account / opportunity | Stage | Owner | Next action | Expected |",
                "|---|---|---|---|---|---:|",
            ]
        )
        for due, _, record in actions:
            marker = "OVERDUE — " if due < as_of else ""
            lines.append(
                "| {due} | {label} | {stage} | {owner} | {marker}{action} | {expected} |".format(
                    due=due.isoformat(),
                    label=row_label(record),
                    stage=clean_text(record.get("stage")),
                    owner=clean_text(record.get("owner")),
                    marker=marker,
                    action=clean_text(record.get("next_action")).replace("|", "/"),
                    expected=money(expected_value(record)),
                )
            )
    lines.extend(
        [
            "",
            "## Human gates",
            "",
            "- Route pricing, promises, proofs, sensitive accounts, and sends for approval.",
            "- Reconcile wins to QuickBooks; do not treat tracker value as accounting truth.",
            "- Stop campaign/outbound volume if same-business-day replies cannot be handled.",
            "",
        ]
    )
    return "\n".join(lines)


def weekly_view(rows: list[dict[str, Any]], as_of: date) -> str:
    start = as_of - timedelta(days=6)
    open_rows = [
        record for record in rows if clean_text(record.get("stage")) in OPEN_STAGES
    ]
    open_pipeline = sum(expected_value(record) for record in open_rows)
    weighted = sum(
        expected_value(record) * effective_probability(record) for record in open_rows
    )
    quoted = [
        record
        for record in rows
        if (quote_date := parse_date(record.get("quote_date"))) and start <= quote_date <= as_of
    ]
    won = [
        record
        for record in rows
        if clean_text(record.get("stage")) == "won"
        and (won_date := parse_date(record.get("order_date")))
        and start <= won_date <= as_of
    ]
    lost = [
        record
        for record in rows
        if clean_text(record.get("stage")) == "lost"
        and (touch_date := parse_date(record.get("last_touch_date")))
        and start <= touch_date <= as_of
    ]
    stages: dict[str, dict[str, float]] = defaultdict(lambda: {"count": 0, "value": 0.0})
    for record in open_rows:
        stage = clean_text(record.get("stage"))
        stages[stage]["count"] += 1
        stages[stage]["value"] += expected_value(record)

    sources: dict[str, dict[str, float]] = defaultdict(lambda: {"count": 0, "value": 0.0})
    for record in open_rows:
        source = clean_text(record.get("lead_source")) or "(blank)"
        sources[source]["count"] += 1
        sources[source]["value"] += expected_value(record)

    loss_counts = Counter(clean_text(record.get("loss_reason")) or "(blank)" for record in lost)
    reorder_pipeline = sum(
        expected_value(record)
        for record in open_rows
        if clean_text(record.get("reorder_flag")) == "yes"
    )
    unreconciled_wins = sum(
        1
        for record in rows
        if clean_text(record.get("stage")) == "won"
        and not clean_text(record.get("qb_invoice_id"))
    )

    lines = [
        f"# Revenue weekly view — {start.isoformat()} to {as_of.isoformat()}",
        "",
        "> Local tracker view. Revenue is reconciled only when QuickBooks IDs are present.",
        "",
        "## Scoreboard",
        "",
        "| Metric | Value |",
        "|---|---:|",
        f"| Open opportunities | {len(open_rows)} |",
        f"| Open pipeline | {money(open_pipeline)} |",
        f"| Weighted pipeline | {money(weighted)} |",
        f"| Reorder pipeline | {money(reorder_pipeline)} |",
        f"| Quotes dated this week | {len(quoted)} |",
        f"| Quoted value this week | {money(sum(parse_number(r.get('quote_amount')) or 0 for r in quoted))} |",
        f"| Orders won this week | {len(won)} |",
        f"| Won value this week | {money(sum(order_value(r) for r in won))} |",
        f"| Unreconciled won rows (all time) | {unreconciled_wins} |",
        "",
        "## Open pipeline by stage",
        "",
        "| Stage | Opportunities | Expected value |",
        "|---|---:|---:|",
    ]
    for stage in STAGE_PROBABILITY:
        if stage not in OPEN_STAGES:
            continue
        values = stages.get(stage, {"count": 0, "value": 0.0})
        lines.append(f"| {stage} | {int(values['count'])} | {money(values['value'])} |")

    lines.extend(
        [
            "",
            "## Open pipeline by lead source",
            "",
            "| Lead source | Opportunities | Expected value |",
            "|---|---:|---:|",
        ]
    )
    for source, values in sorted(
        sources.items(), key=lambda item: (-item[1]["value"], item[0])
    ):
        lines.append(f"| {source} | {int(values['count'])} | {money(values['value'])} |")

    lines.extend(["", "## Losses closed this week", ""])
    if not loss_counts:
        lines.append("_No losses dated by last touch this week._")
    else:
        lines.extend(["| Reason | Count |", "|---|---:|"])
        for reason, count in loss_counts.most_common():
            lines.append(f"| {reason} | {count} |")

    lines.extend(
        [
            "",
            "## Review decisions",
            "",
            "1. Which source/offer produced the strongest qualified movement?",
            "2. Which overdue next actions or quote blockers need Kenny/Maclaine?",
            "3. Which one experiment changes next week?",
            "4. Is operational capacity sufficient before more demand is created?",
            "",
        ]
    )
    return "\n".join(lines)


def render_issue_summary(path: Path, rows: list[dict[str, Any]], issues: list[Issue]) -> str:
    errors = [issue for issue in issues if issue.severity == "error"]
    warnings = [issue for issue in issues if issue.severity == "warning"]
    payload = {
        "workbook": str(path),
        "data_rows": len(rows),
        "errors": len(errors),
        "warnings": len(warnings),
        "issues": [asdict(issue) for issue in issues],
    }
    return json.dumps(payload, indent=2)


def command_validate(args: argparse.Namespace) -> int:
    path = Path(args.workbook)
    rows, issues = validate_workbook(path)
    print(render_issue_summary(path, rows, issues))
    return 1 if any(issue.severity == "error" for issue in issues) else 0


def command_views(args: argparse.Namespace) -> int:
    path = Path(args.workbook)
    rows, issues = validate_workbook(path)
    errors = [issue for issue in issues if issue.severity == "error"]
    print(render_issue_summary(path, rows, issues))
    if errors:
        print("Views were not written because validation has errors.")
        return 1

    as_of = date.fromisoformat(args.as_of) if args.as_of else date.today()
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    daily_path = output_dir / f"revenue-daily-{as_of.isoformat()}.md"
    weekly_path = output_dir / f"revenue-weekly-{as_of.isoformat()}.md"
    daily_path.write_text(daily_view(rows, as_of) + "\n", encoding="utf-8")
    weekly_path.write_text(weekly_view(rows, as_of) + "\n", encoding="utf-8")
    print(f"Wrote {daily_path}")
    print(f"Wrote {weekly_path}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)

    validate = subparsers.add_parser("validate", help="Validate workbook tracker rows")
    validate.add_argument("--workbook", required=True)
    validate.set_defaults(func=command_validate)

    views = subparsers.add_parser("views", help="Validate and write daily/weekly views")
    views.add_argument("--workbook", required=True)
    views.add_argument("--output-dir", required=True)
    views.add_argument("--as-of", help="ISO date; defaults to today")
    views.set_defaults(func=command_views)
    return parser


def main() -> int:
    args = build_parser().parse_args()
    return int(args.func(args))


if __name__ == "__main__":
    raise SystemExit(main())
