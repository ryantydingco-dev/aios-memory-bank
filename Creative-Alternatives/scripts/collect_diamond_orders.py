"""
AIOS Client Workspace — Diamond Open Orders Collector

Pulls the shared production-schedule Google Sheet that CA and Diamond
(contract printer — apparel/embroidery) both work from, and loads the OPEN
tab into the database. Link-shared like the Viking sheet, so no API key —
we download the workbook as xlsx and parse it with openpyxl.

Sheet: "DIAMOND <> CREATIVE ALTERNATIVES OPEN ORDERS"

Sheet layout (OPEN tab):
    (Dates/rush) | PO Number | Company | Description | Ship To Date |
    Event? | ART Approved? | Shipping Method | Tracking Number | Notes

Rows come in weekly blocks ("July 13-17, 2026" header rows in column A).
Column A on an order row carries rush flags ("MUST THURSDAY"). BOLD rows
are hard ship dates — captured as must_ship_bold. There is no Status
column; a tracking number means the order shipped (rows then move to the
CLOSED tab, which has a different layout and is not collected).

Table is fully reloaded on every run — the sheet is the source of truth.
"""

import io
import urllib.request
from datetime import datetime, timezone

SHEET_ID = "1-pcbgUX1fMbKrmpTFKWW15id_D0xNP93ArovtYvg2_4"
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
TAB_NAME = "OPEN"


def _text(v):
    """Cell value -> stripped string or None. Floats like 27598.0 -> '27598'."""
    if v is None:
        return None
    if isinstance(v, float) and v == int(v):
        v = int(v)
    s = str(v).strip()
    return s or None


def _parse_date(v):
    """Return (iso_date_or_None, raw_string). Tolerates typo years and free text."""
    if v is None:
        return None, None
    if isinstance(v, datetime):
        raw = v.strftime("%Y-%m-%d")
        if 2020 <= v.year <= 2035:
            return raw, raw
        return None, raw  # typo year — keep raw, don't sort on it
    raw = str(v).strip()
    if not raw:
        return None, None
    for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            d = datetime.strptime(raw, fmt)
            if 2020 <= d.year <= 2035:
                return d.strftime("%Y-%m-%d"), raw
        except ValueError:
            pass
    return None, raw


def collect():
    """Download the workbook and parse the OPEN tab."""
    try:
        req = urllib.request.Request(EXPORT_URL, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            payload = resp.read()
    except Exception as e:
        return {"source": "diamond_orders", "status": "skipped",
                "reason": f"sheet download failed ({e})"}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(payload))
        if TAB_NAME not in wb.sheetnames:
            return {"source": "diamond_orders", "status": "error",
                    "reason": f"tab '{TAB_NAME}' not found (tabs: {wb.sheetnames})"}
        ws = wb[TAB_NAME]

        orders = []
        week_label = None
        for row in ws.iter_rows(min_row=1, max_col=10):
            vals = [c.value for c in row]
            a = _text(vals[0])
            body = [_text(v) for v in vals[1:4]]  # PO, Company, Description

            if body[0] == "PO Number":
                continue
            if not any(body):
                if a:
                    week_label = a  # e.g. "July 13-17, 2026"
                continue

            ship_to, ship_to_raw = _parse_date(vals[4])
            company_cell, desc_cell = row[2], row[3]
            bold = bool((desc_cell.font and desc_cell.font.bold)
                        or (company_cell.font and company_cell.font.bold))
            orders.append({
                "week_label": week_label,
                "rush_note": a,
                "po_number": body[0], "company": body[1], "description": body[2],
                "ship_to": ship_to, "ship_to_raw": ship_to_raw,
                "event": _text(vals[5]), "art_status": _text(vals[6]),
                "ship_method": _text(vals[7]), "tracking": _text(vals[8]),
                "notes": _text(vals[9]),
                "must_ship_bold": 1 if bold else 0,
            })

        return {"source": "diamond_orders", "status": "success",
                "data": {"orders": orders}}
    except Exception as e:
        return {"source": "diamond_orders", "status": "error", "reason": str(e)}


def write(conn, result, date):
    """Full reload of diamond_open_orders. Returns records written."""
    conn.execute("""
        CREATE TABLE IF NOT EXISTS diamond_open_orders (
            date TEXT NOT NULL,            -- collection date
            week_label TEXT,
            rush_note TEXT,
            po_number TEXT,
            company TEXT,
            description TEXT,
            ship_to TEXT,                  -- ISO date when parseable
            ship_to_raw TEXT,              -- what the sheet actually says
            event TEXT,
            art_status TEXT,
            ship_method TEXT,
            tracking TEXT,
            notes TEXT,
            must_ship_bold INTEGER,
            collected_at TEXT
        )
    """)

    if result.get("status") != "success":
        conn.commit()
        return 0

    conn.execute("DELETE FROM diamond_open_orders")
    collected_at = datetime.now(timezone.utc).isoformat()
    records = 0
    for r in result["data"]["orders"]:
        conn.execute(
            "INSERT INTO diamond_open_orders (date, week_label, rush_note, po_number, "
            "company, description, ship_to, ship_to_raw, event, art_status, "
            "ship_method, tracking, notes, must_ship_bold, collected_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (date, r["week_label"], r["rush_note"], r["po_number"], r["company"],
             r["description"], r["ship_to"], r["ship_to_raw"], r["event"],
             r["art_status"], r["ship_method"], r["tracking"], r["notes"],
             r["must_ship_bold"], collected_at))
        records += 1
    conn.commit()
    return records


if __name__ == "__main__":
    result = collect()
    if result["status"] == "success":
        orders = result["data"]["orders"]
        print(f"{len(orders)} open orders")
        for o in orders:
            flag = " [BOLD]" if o["must_ship_bold"] else ""
            rush = f" [{o['rush_note']}]" if o["rush_note"] else ""
            print(f"  {o['company']} | {o['description']} | ship {o['ship_to_raw']} | "
                  f"{o['art_status'] or 'no art status'}{flag}{rush}")
    else:
        print(f"{result['status']}: {result.get('reason', '')}")
