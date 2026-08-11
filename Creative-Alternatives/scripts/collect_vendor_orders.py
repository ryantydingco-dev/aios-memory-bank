"""
AIOS Client Workspace — Random Vendors Open Orders Collector

Pulls the shared Google Sheet tracking CA's promo-product POs across
one-off vendors (Wow Line, NC Custom, PrimeLine, Koozie Group, …) and
loads the OPEN tab into the database. Link-shared like the Viking sheet,
so no API key — we download the workbook as xlsx and parse it with openpyxl.

Sheet: "RANDOM VENDORS <> CREATIVE ALTERNATIVES"

Sheet layout (OPEN tab — note the extra leading column vs the printer sheets):
    (misc) | Dates | PO Number | Vendor | Customer | Description |
    ART Approved? | In Hand Date | Misc | Shipping Method | Status |
    Tracking Number | Notes

Rows come in weekly blocks ("July 13-July 17" header rows in column B,
plus an "Older Orders" block). The In Hand Date is when the CUSTOMER must
have the goods (vendors drop-ship), so it is the deadline we track. BOLD
rows are hard dates — captured as must_ship_bold. The SHIPPED tab has a
different layout and is not collected.

Table is fully reloaded on every run — the sheet is the source of truth.
"""

import io
import urllib.request
from datetime import datetime, timezone

SHEET_ID = "1EGD_i02NDybDQw8tdPPcxrUl6ANhW-hdg1cEQnC3s9o"
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
TAB_NAME = "OPEN"


def _text(v):
    """Cell value -> stripped string or None. Floats like 27598.0 -> '27598'."""
    if v is None:
        return None
    if isinstance(v, float) and v == int(v):
        v = int(v)
    s = str(v).strip()
    return s or None


def _parse_date(v):
    """Return (iso_date_or_None, raw_string). Tolerates typo years and free text."""
    if v is None:
        return None, None
    if isinstance(v, datetime):
        raw = v.strftime("%Y-%m-%d")
        if 2020 <= v.year <= 2035:
            return raw, raw
        return None, raw  # typo year — keep raw, don't sort on it
    raw = str(v).strip()
    if not raw:
        return None, None
    for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            d = datetime.strptime(raw, fmt)
            if 2020 <= d.year <= 2035:
                return d.strftime("%Y-%m-%d"), raw
        except ValueError:
            pass
    return None, raw


def collect():
    """Download the workbook and parse the OPEN tab."""
    try:
        req = urllib.request.Request(EXPORT_URL, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            payload = resp.read()
    except Exception as e:
        return {"source": "vendor_orders", "status": "skipped",
                "reason": f"sheet download failed ({e})"}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(payload))
        if TAB_NAME not in wb.sheetnames:
            return {"source": "vendor_orders", "status": "error",
                    "reason": f"tab '{TAB_NAME}' not found (tabs: {wb.sheetnames})"}
        ws = wb[TAB_NAME]

        orders = []
        week_label = None
        for row in ws.iter_rows(min_row=1, max_col=13):
            vals = [c.value for c in row]
            rush = _text(vals[0])
            b = _text(vals[1])
            body = [_text(v) for v in vals[2:6]]  # PO, Vendor, Customer, Description

            if body[0] == "PO Number":
                continue
            if not any(body):
                if b:
                    week_label = b  # e.g. "July 13-July 17" or "Older Orders"
                continue

            in_hand, in_hand_raw = _parse_date(vals[7])
            customer_cell, desc_cell = row[4], row[5]
            bold = bool((desc_cell.font and desc_cell.font.bold)
                        or (customer_cell.font and customer_cell.font.bold))
            orders.append({
                "week_label": week_label,
                "rush_note": rush or b,   # rush flags land in A or on the B cell
                "po_number": body[0], "vendor": body[1],
                "customer": body[2], "description": body[3],
                "art_status": _text(vals[6]),
                "in_hand": in_hand, "in_hand_raw": in_hand_raw,
                "misc": _text(vals[8]), "ship_method": _text(vals[9]),
                "status": _text(vals[10]), "tracking": _text(vals[11]),
                "notes": _text(vals[12]),
                "must_ship_bold": 1 if bold else 0,
            })

        return {"source": "vendor_orders", "status": "success",
                "data": {"orders": orders}}
    except Exception as e:
        return {"source": "vendor_orders", "status": "error", "reason": str(e)}


def write(conn, result, date):
    """Full reload of vendor_open_orders. Returns records written."""
    conn.execute("""
        CREATE TABLE IF NOT EXISTS vendor_open_orders (
            date TEXT NOT NULL,            -- collection date
            week_label TEXT,
            rush_note TEXT,
            po_number TEXT,
            vendor TEXT,
            customer TEXT,
            description TEXT,
            art_status TEXT,
            in_hand TEXT,                  -- ISO date when parseable
            in_hand_raw TEXT,              -- what the sheet actually says
            misc TEXT,
            ship_method TEXT,
            status TEXT,
            tracking TEXT,
            notes TEXT,
            must_ship_bold INTEGER,
            collected_at TEXT
        )
    """)

    if result.get("status") != "success":
        conn.commit()
        return 0

    conn.execute("DELETE FROM vendor_open_orders")
    collected_at = datetime.now(timezone.utc).isoformat()
    records = 0
    for r in result["data"]["orders"]:
        conn.execute(
            "INSERT INTO vendor_open_orders (date, week_label, rush_note, po_number, "
            "vendor, customer, description, art_status, in_hand, in_hand_raw, misc, "
            "ship_method, status, tracking, notes, must_ship_bold, collected_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (date, r["week_label"], r["rush_note"], r["po_number"], r["vendor"],
             r["customer"], r["description"], r["art_status"], r["in_hand"],
             r["in_hand_raw"], r["misc"], r["ship_method"], r["status"],
             r["tracking"], r["notes"], r["must_ship_bold"], collected_at))
        records += 1
    conn.commit()
    return records


if __name__ == "__main__":
    result = collect()
    if result["status"] == "success":
        orders = result["data"]["orders"]
        print(f"{len(orders)} open orders")
        for o in orders:
            flag = " [BOLD]" if o["must_ship_bold"] else ""
            print(f"  {o['customer']} | {o['description']} (via {o['vendor']}) | "
                  f"in hand {o['in_hand_raw']} | {o['status'] or o['art_status'] or 'no status'}{flag}")
    else:
        print(f"{result['status']}: {result.get('reason', '')}")
