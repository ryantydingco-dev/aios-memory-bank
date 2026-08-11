"""
AIOS Client Workspace — Viking Open Orders Collector

Pulls the shared production-schedule Google Sheet that CA and Viking
(contract printer, Lindenhurst) both work from, and loads the OPEN tab
into the database. The sheet is link-shared, so no API key is needed —
we download the workbook as xlsx and parse it with openpyxl.

Sheet layout (OPEN tab):
    Dates | PO Number | Viking Number | Department | Company | Description |
    NEED To Ship By Date | EVENT | Shipping Method | Status | Tracking Number |
    Approved? | Notes

Rows come in weekly blocks ("July 13-July 17" header rows). The Dates column
on an order row carries rush flags ("NEW RUSH!!!!!!!", "MUST MONDAY").
Per the sheet's own reminder, BOLD rows must go out on their ship date —
we capture bold as must_ship_bold.

The CLOSED tab uses a different column order and has no header row, so it
is not collected (revisit if we want shipped-order history).

Table is fully reloaded on every run — the sheet is the source of truth.
"""

import io
import tempfile
import urllib.request
from datetime import datetime, timezone

SHEET_ID = "1baN49Mv4CAg40yPjXwSo20doOpekNWqzfeqrcIXla10"
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
TAB_NAME = "OPEN"


def _text(v):
    """Cell value -> stripped string or None. Floats like 27598.0 -> '27598'."""
    if v is None:
        return None
    if isinstance(v, float) and v == int(v):
        v = int(v)
    s = str(v).strip()
    return s or None


def _parse_ship_by(v):
    """Return (iso_date_or_None, raw_string). Tolerates typos like year 2426
    and free text like '04/23 or 04/24'."""
    if v is None:
        return None, None
    if isinstance(v, datetime):
        raw = v.strftime("%Y-%m-%d")
        if 2020 <= v.year <= 2035:
            return raw, raw
        return None, raw  # typo year — keep raw, don't sort on it
    raw = str(v).strip()
    if not raw:
        return None, None
    for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            d = datetime.strptime(raw, fmt)
            if 2020 <= d.year <= 2035:
                return d.strftime("%Y-%m-%d"), raw
        except ValueError:
            pass
    return None, raw


def collect():
    """Download the workbook and parse the OPEN tab."""
    try:
        req = urllib.request.Request(EXPORT_URL, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            payload = resp.read()
    except Exception as e:
        return {"source": "viking_orders", "status": "skipped",
                "reason": f"sheet download failed ({e})"}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(payload))
        if TAB_NAME not in wb.sheetnames:
            return {"source": "viking_orders", "status": "error",
                    "reason": f"tab '{TAB_NAME}' not found (tabs: {wb.sheetnames})"}
        ws = wb[TAB_NAME]

        orders = []
        week_label = None
        for row in ws.iter_rows(min_row=1, max_col=13):
            vals = [c.value for c in row]
            a = _text(vals[0])
            body = [_text(v) for v in vals[1:6]]  # PO, Viking#, Dept, Company, Description

            if a and "REMINDER" in a.upper():
                continue
            if a == "Dates" and body[0] == "PO Number":
                continue
            if not any(body):
                if a:
                    week_label = a  # e.g. "July 13-July 17"
                continue

            ship_by, ship_by_raw = _parse_ship_by(vals[6])
            desc_cell, company_cell = row[5], row[4]
            bold = bool((desc_cell.font and desc_cell.font.bold)
                        or (company_cell.font and company_cell.font.bold))
            orders.append({
                "week_label": week_label,
                "rush_note": a,
                "po_number": body[0], "viking_number": body[1],
                "department": body[2], "company": body[3], "description": body[4],
                "ship_by": ship_by, "ship_by_raw": ship_by_raw,
                "event": _text(vals[7]), "ship_method": _text(vals[8]),
                "status": _text(vals[9]), "tracking": _text(vals[10]),
                "approved": _text(vals[11]), "notes": _text(vals[12]),
                "must_ship_bold": 1 if bold else 0,
            })

        return {"source": "viking_orders", "status": "success",
                "data": {"orders": orders}}
    except Exception as e:
        return {"source": "viking_orders", "status": "error", "reason": str(e)}


def write(conn, result, date):
    """Full reload of viking_open_orders. Returns records written."""
    conn.execute("""
        CREATE TABLE IF NOT EXISTS viking_open_orders (
            date TEXT NOT NULL,            -- collection date
            week_label TEXT,
            rush_note TEXT,
            po_number TEXT,
            viking_number TEXT,
            department TEXT,
            company TEXT,
            description TEXT,
            ship_by TEXT,                  -- ISO date when parseable
            ship_by_raw TEXT,              -- what the sheet actually says
            event TEXT,
            ship_method TEXT,
            status TEXT,
            tracking TEXT,
            approved TEXT,
            notes TEXT,
            must_ship_bold INTEGER,
            collected_at TEXT
        )
    """)

    if result.get("status") != "success":
        conn.commit()
        return 0

    conn.execute("DELETE FROM viking_open_orders")
    collected_at = datetime.now(timezone.utc).isoformat()
    records = 0
    for r in result["data"]["orders"]:
        conn.execute(
            "INSERT INTO viking_open_orders (date, week_label, rush_note, po_number, "
            "viking_number, department, company, description, ship_by, ship_by_raw, "
            "event, ship_method, status, tracking, approved, notes, must_ship_bold, "
            "collected_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (date, r["week_label"], r["rush_note"], r["po_number"], r["viking_number"],
             r["department"], r["company"], r["description"], r["ship_by"],
             r["ship_by_raw"], r["event"], r["ship_method"], r["status"],
             r["tracking"], r["approved"], r["notes"], r["must_ship_bold"],
             collected_at))
        records += 1
    conn.commit()
    return records


if __name__ == "__main__":
    result = collect()
    if result["status"] == "success":
        orders = result["data"]["orders"]
        print(f"{len(orders)} open orders")
        for o in orders:
            flag = " [BOLD]" if o["must_ship_bold"] else ""
            print(f"  {o['company']} | {o['description']} | ship {o['ship_by_raw']} | "
                  f"{o['status'] or 'NO STATUS'}{flag}")
    else:
        print(f"{result['status']}: {result.get('reason', '')}")
